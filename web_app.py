from flask import Flask, request, render_template, jsonify
import os
from openpyxl import load_workbook

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/save', methods=['POST'])
def save_data():
    data = request.json
    file_type = data['type']
    item = data['item']
    filename = os.path.join(BASE_DIR, f'{file_type}.xlsx')

    wb = load_workbook(filename) if os.path.exists(filename) else None
    if not wb:
        wb = Workbook()
        ws = wb.active
        ws.append(['id', 'title', 'description', 'deadline'] if file_type == 'tenders' else
                  ['id', 'title', 'text'] if file_type == 'news' else ['id', 'name', 'description'])
    else:
        ws = wb.active

    # اضافه کردن یا به‌روزرسانی ردیف
    for i, row in enumerate(ws.rows, 1):
        if i > 1 and (row[1].value == item.get('id', i)):
            for j, key in enumerate(['title', 'description', 'deadline'] if file_type == 'tenders' else
                                   ['title', 'text'] if file_type == 'news' else ['name', 'description'], 2):
                ws.cell(row=i, column=j, value=item.get(key))
            break
    else:
        ws.append([len(list(ws.rows)), item['title'], item.get('description', ''), item.get('deadline', '')])
    wb.save(filename)
    return jsonify({"status": "success"})

@app.route('/api/load/<file_type>')
def load_data(file_type):
    filename = os.path.join(BASE_DIR, f'{file_type}.xlsx')
    if not os.path.exists(filename):
        return jsonify([])
    wb = load_workbook(filename)
    ws = wb.active
    return jsonify([[cell.value for cell in row] for row in ws.rows if any(cell.value for cell in row)])

if __name__ == '__main__':
    app.run(debug=True)